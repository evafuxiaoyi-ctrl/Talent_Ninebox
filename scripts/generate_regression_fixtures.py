from __future__ import annotations

import argparse
import shutil
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from talent_ninebox.core.models import ProcessOptions
from talent_ninebox.core.processor import process_zip

HEADERS = [
    "序号",
    "工号",
    "姓名",
    "一级部门",
    "二级部门",
    "所属部门",
    "初始九宫格落位",
    "校准后九宫格落位",
    "价值观综合得分",
    "离职风险",
    "备注公式",
]

PLACEMENTS = [
    "高潜力 / 低绩效",
    "高潜力 / 中绩效",
    "高潜力 / 高绩效",
    "中潜力 / 低绩效",
    "中潜力 / 中绩效",
    "中潜力 / 高绩效",
    "低潜力 / 低绩效",
    "低潜力 / 中绩效",
    "低潜力 / 高绩效",
]


def _row(person_index: int, placement: str, *, blank_initial: bool = False) -> list[object]:
    value_scores = [4.5, 3.5, 2, 4, 1.5, 3, 4.2, 2, 3.8]
    risks = ["低", "中", "高", "低", "高", "中", "低", "中", "高"]
    placement_index = (person_index - 1) % len(PLACEMENTS)
    return [
        person_index,
        f"E{person_index:04d}",
        f"测试人员{person_index:03d}",
        f"一级部门{(person_index - 1) % 3 + 1}",
        f"二级部门{(person_index - 1) % 5 + 1}",
        f"所属部门{(person_index - 1) % 4 + 1}",
        "" if blank_initial else placement,
        PLACEMENTS[(placement_index + 1) % len(PLACEMENTS)],
        value_scores[placement_index],
        risks[placement_index],
        None,
    ]


def make_workbook(path: Path, start_index: int, count: int, *, missing_initial: bool = False, missing_final: bool = False, blank_initial_rows: set[int] | None = None) -> None:
    blank_initial_rows = blank_initial_rows or set()
    wb = Workbook()
    ws = wb.active
    ws.title = "人才盘点"
    ws["A1"] = "人才盘点信息收集表"
    headers = [
        header
        for header in HEADERS
        if not (missing_initial and header == "初始九宫格落位")
        and not (missing_final and header == "校准后九宫格落位")
    ]
    ws.append(headers)

    for offset in range(count):
        person_index = start_index + offset
        placement = PLACEMENTS[(person_index - 1) % len(PLACEMENTS)]
        row = _row(person_index, placement, blank_initial=person_index in blank_initial_rows)
        if missing_initial:
            row.pop(6)
        if missing_final:
            row.pop(6 if missing_initial else 7)
        ws.append(row)
        excel_row = ws.max_row
        formula_col = len(headers)
        ws.cell(excel_row, formula_col).value = f'=IF(C{excel_row}="","",C{excel_row}&"-"&D{excel_row})'

    params = wb.create_sheet("参数")
    params.append(["离职风险选项", "价值观阈值"])
    params.append(["高", 4])
    params.append(["中", 2])
    params.append(["低", None])

    risk_col = headers.index("离职风险") + 1
    risk_letter = ws.cell(1, risk_col).column_letter
    validation = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
    validation.add(f"{risk_letter}3:{risk_letter}{ws.max_row}")
    ws.add_data_validation(validation)
    ws.freeze_panes = "A3"
    wb.save(path)


def make_zip(
    zip_path: Path,
    file_count: int,
    *,
    rows_per_file: int = 3,
    missing_initial_file: int | None = None,
    missing_final_file: int | None = None,
    blank_initial_file: int | None = None,
) -> None:
    work_dir = zip_path.parent / f"_{zip_path.stem}_files"
    if work_dir.exists():
        shutil.rmtree(work_dir)
    work_dir.mkdir(parents=True)
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file_index in range(1, file_count + 1):
            path = work_dir / f"部门{file_index:02d}_人才盘点.xlsx"
            start_index = (file_index - 1) * rows_per_file + 1
            blank_rows = {start_index} if blank_initial_file == file_index else set()
            make_workbook(
                path,
                start_index,
                rows_per_file,
                missing_initial=missing_initial_file == file_index,
                missing_final=missing_final_file == file_index,
                blank_initial_rows=blank_rows,
            )
            archive.write(path, f"部门{file_index:02d}/{path.name}")
        archive.writestr("__MACOSX/ignored.xlsx", "ignored")
        archive.writestr("~$temp.xlsx", "ignored")


def generate_fixture_set(output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "initial_10": output_dir / "初版_10个正常文件.zip",
        "initial_30": output_dir / "初版_30个正常文件.zip",
        "missing_field": output_dir / "异常_字段缺失.zip",
        "blank_placement": output_dir / "异常_落位为空.zip",
        "value_risk": output_dir / "覆盖_价值观和离职风险.zip",
        "final_workbook": output_dir / "终版_整合后文件.xlsx",
    }
    make_zip(paths["initial_10"], 10)
    make_zip(paths["initial_30"], 30)
    make_zip(paths["missing_field"], 3, missing_initial_file=2)
    make_zip(paths["blank_placement"], 3, blank_initial_file=2)
    make_zip(paths["value_risk"], 9, rows_per_file=1)

    initial_result = process_zip(
        paths["initial_10"],
        output_dir / "_generated_outputs",
        ProcessOptions(placement_mode="initial", stage_label="初版整合"),
    )
    workbook = load_workbook(initial_result.output_file)
    workbook.save(paths["final_workbook"])
    return paths


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate Talent Ninebox regression Excel fixtures.")
    parser.add_argument("--output-dir", default="tmp/regression-fixtures")
    args = parser.parse_args()
    paths = generate_fixture_set(Path(args.output_dir))
    for name, path in paths.items():
        print(f"{name}: {path}")


if __name__ == "__main__":
    main()
