from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import Issue, ProcessSummary

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_table(ws: Worksheet, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)


def write_summary(ws: Worksheet, summary: ProcessSummary) -> None:
    rows: list[list[object]] = []
    for key, value in summary.as_dict().items():
        if isinstance(value, dict):
            rows.append([key, "", ""])
            for sub_key, sub_value in value.items():
                ratio = sub_value / summary.total_people if summary.total_people else 0
                rows.append([f"  {sub_key}", sub_value, ratio])
        else:
            rows.append([key, value, ""])
    write_table(ws, ["指标", "值", "比例"], rows)
    for cell in ws["C"]:
        if isinstance(cell.value, float):
            cell.number_format = "0.00%"


def write_issues(ws: Worksheet, issues: list[Issue]) -> None:
    rows = [
        [i.level, i.category, i.source_file, i.sheet, i.row or "", i.field, i.raw_value, i.message]
        for i in issues
    ]
    write_table(ws, ["级别", "类型", "来源文件", "Sheet", "行号", "字段", "原始值", "问题说明"], rows)


def write_instructions(ws: Worksheet) -> None:
    rows = [
        ["整合总表", "合并后的人员数据，保留主模板字段顺序，并附加来源辅助列。"],
        ["人才九宫格", "单元格版九宫格，姓名文本可直接编辑、复制到 PPT 或截图。"],
        ["异常报告", "查看文件、模板、数据和落位识别异常。"],
        ["公式说明", "系统保留公式并尽量平移同 Sheet 相对引用；公式结果由 Excel 打开后刷新。"],
    ]
    write_table(ws, ["Sheet", "说明"], rows)
