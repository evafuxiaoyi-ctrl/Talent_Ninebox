from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .field_utils import has_field, has_placement_field
from .models import Issue, ProcessOptions, WorkbookInfo

def normalize_header(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _placement_label(options: ProcessOptions) -> str:
    if options.placement_mode == "initial":
        return "初始九宫格落位"
    if options.placement_mode == "final":
        return "校准后九宫格落位"
    return "人才九宫格落位"


def inspect_workbook(source_name: str, path: Path, options: ProcessOptions) -> tuple[WorkbookInfo | None, list[Issue]]:
    issues: list[Issue] = []
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as exc:
        return None, [
            Issue(
                "error",
                "文件读取异常",
                "文件无法读取，可能已损坏、加密，或并非有效的 .xlsx 文件。",
                source_file=source_name,
                raw_value=type(exc).__name__,
            )
        ]

    for ws in wb.worksheets:
        max_row = min(options.max_header_scan_rows, ws.max_row)
        for row_idx in range(1, max_row + 1):
            values = [normalize_header(cell.value) for cell in ws[row_idx]]
            if has_field(values, "姓名") and has_placement_field(values, options.placement_mode):
                headers = values
                while headers and headers[-1] == "":
                    headers.pop()
                normalized = [normalize_header(v) for v in headers]
                return WorkbookInfo(source_name, path, ws.title, row_idx, headers, normalized), issues

    return None, [
        Issue(
            "error",
            "模板异常",
            f"前 {options.max_header_scan_rows} 行未找到同时包含「姓名/员工姓名」和「{_placement_label(options)}」的表头行。",
            source_file=source_name,
        )
    ]
