from __future__ import annotations

import re
import shutil
import tempfile
import zipfile
import copy
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

HEADER_ROW = 3
DATA_START_ROW = HEADER_ROW + 1
MAX_SPLIT_GROUPS = 80
REFERENCE_SHEET_NAMES = {"参数"}
ALLOWED_SPLIT_FIELDS = [
    "盘点人员类型",
    "一级部门",
    "二级部门",
    "三级部门",
    "盘点人",
]
SAMPLE_ROW_MARKERS = {"示例行"}
WORKBOOK_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


@dataclass(frozen=True)
class SplitField:
    name: str
    column_index: int
    column_letter: str


@dataclass(frozen=True)
class SplitSheet:
    name: str
    fields: list[SplitField]


@dataclass(frozen=True)
class SplitResult:
    output_file: Path
    summary: dict[str, object]


def _clean_header(value: object) -> str:
    return str(value or "").strip()


def _safe_filename_part(value: object) -> str:
    text = str(value or "").strip() or "未填写"
    text = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text[:60] or "未填写"


def _is_sample_row(sheet, row: int) -> bool:
    return _clean_header(sheet.cell(row=row, column=1).value) in SAMPLE_ROW_MARKERS


def _iter_person_data_rows(sheet):
    for row in range(DATA_START_ROW, sheet.max_row + 1):
        if _is_sample_row(sheet, row):
            continue
        yield row


def _fields_for_sheet(sheet) -> list[SplitField]:
    fields: list[SplitField] = []
    seen: set[str] = set()
    header_rows = list(sheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))
    if not header_rows:
        return fields
    for cell in header_rows[0]:
        header = _clean_header(cell.value)
        if header in ALLOWED_SPLIT_FIELDS and header not in seen:
            fields.append(SplitField(header, cell.column, get_column_letter(cell.column)))
            seen.add(header)
    return fields


def list_split_sheets(workbook_path: Path) -> list[SplitSheet]:
    try:
        workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    except Exception as exc:
        raise ValueError("文件无法读取，可能已损坏、加密，或并非有效的 .xlsx 文件。") from exc
    try:
        sheets = [
            SplitSheet(sheet.title, fields)
            for sheet in workbook.worksheets
            if sheet.title not in REFERENCE_SHEET_NAMES and (fields := _fields_for_sheet(sheet))
        ]
        if not sheets:
            raise ValueError("所有 Sheet 的第 3 行都未找到可用于拆分的字段，请确认模板包含盘点人员类型、一级部门、二级部门、三级部门或盘点人。")
        return sheets
    finally:
        workbook.close()


def list_split_fields(workbook_path: Path, sheet_name: str | None = None) -> list[SplitField]:
    sheets = list_split_sheets(workbook_path)
    if sheet_name is None:
        return sheets[0].fields
    for sheet in sheets:
        if sheet.name == sheet_name:
            return sheet.fields
    raise ValueError(f"未找到可用于拆分的 Sheet「{sheet_name}」。")


def _group_values(workbook_path: Path, sheet_name: str, field_name: str) -> tuple[int, dict[str, int]]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"未找到 Sheet「{sheet_name}」。")
        sheet = workbook[sheet_name]
        field_column = None
        data_columns: list[int] = []
        for cell in sheet[HEADER_ROW]:
            header = _clean_header(cell.value)
            if header in ALLOWED_SPLIT_FIELDS:
                data_columns.append(cell.column)
            if header == field_name:
                field_column = cell.column
        if field_column is None:
            raise ValueError(f"第 3 行未找到拆分字段「{field_name}」。")
        if not data_columns:
            raise ValueError(f"Sheet「{sheet_name}」第 3 行未找到可用于判断数据行的字段。")

        groups: dict[str, int] = {}
        for row in _iter_person_data_rows(sheet):
            row_values = [sheet.cell(row=row, column=column).value for column in data_columns]
            if not any(value not in (None, "") for value in row_values):
                continue
            value = str(sheet.cell(row=row, column=field_column).value or "").strip() or "未填写"
            groups[value] = groups.get(value, 0) + 1

        if not groups:
            raise ValueError("未找到可拆分的数据行，请确认第 4 行开始存在人员数据。")
        if len(groups) > MAX_SPLIT_GROUPS:
            raise ValueError(f"拆分后将生成 {len(groups)} 个文件，超过 {MAX_SPLIT_GROUPS} 个上限，请更换为更上一级部门字段。")
        return field_column, groups
    finally:
        workbook.close()


def _is_formula(value: object) -> bool:
    return (isinstance(value, str) and value.startswith("=")) or isinstance(value, ArrayFormula)


def _translate_formula(value: object, source, target) -> object:
    if isinstance(value, ArrayFormula):
        origin = value.ref or source.coordinate
        text = value.text
    else:
        origin = source.coordinate
        text = value
    if not isinstance(text, str):
        return value
    if source.coordinate != target.coordinate:
        try:
            text = Translator(text, origin=origin).translate_formula(target.coordinate)
        except Exception:
            pass
    if isinstance(value, ArrayFormula):
        return ArrayFormula(target.coordinate, text)
    return text


def _detect_formula_templates(sheet) -> dict[int, tuple[object, int]]:
    templates: dict[int, tuple[object, int]] = {}
    for row in range(DATA_START_ROW, min(sheet.max_row, DATA_START_ROW + 30) + 1):
        for cell in sheet[row]:
            if cell.column in templates:
                continue
            if _is_formula(cell.value):
                templates[cell.column] = (copy.copy(cell.value), row)
    return templates


def _sheet_xml_path(workbook_path: Path, sheet_name: str) -> str | None:
    with zipfile.ZipFile(workbook_path) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    sheet_rel_id = None
    for sheet in workbook_root.findall(f".//{{{WORKBOOK_NS}}}sheet"):
        if sheet.attrib.get("name") == sheet_name:
            sheet_rel_id = sheet.attrib.get(f"{{{REL_NS}}}id")
            break
    if sheet_rel_id is None:
        return None

    for rel in rels_root.findall(f"{{{PACKAGE_REL_NS}}}Relationship"):
        if rel.attrib.get("Id") != sheet_rel_id:
            continue
        target = rel.attrib.get("Target", "")
        if target.startswith("/"):
            target = target.lstrip("/")
        elif not target.startswith("xl/"):
            target = f"xl/{target}"
        return target
    return None


def _worksheet_opening_tag(xml: str) -> str | None:
    match = re.search(r"<worksheet\b[^>]*>", xml)
    return match.group(0) if match else None


def _worksheet_ext_list(xml: str) -> str | None:
    match = re.search(r"<extLst\b.*?</extLst>", xml, flags=re.DOTALL)
    return match.group(0) if match else None


def _x14_data_validation_fragments(xml: str) -> list[str]:
    fragments: list[str] = []
    allowed_attrs = {"type", "allowBlank", "showDropDown", "showInputMessage", "showErrorMessage", "errorStyle", "operator"}
    for match in re.finditer(r"<x14:dataValidation\b([^>]*)>.*?</x14:dataValidation>", xml, flags=re.DOTALL):
        attrs = dict(re.findall(r"([\w:]+)=\"([^\"]*)\"", match.group(1)))
        sqref_match = re.search(r"<xm:sqref>(.*?)</xm:sqref>", match.group(0), flags=re.DOTALL)
        formula_match = re.search(r"<xm:f>(.*?)</xm:f>", match.group(0), flags=re.DOTALL)
        if not sqref_match or not formula_match:
            continue

        normal_attrs = []
        for name, value in attrs.items():
            if ":" in name or name not in allowed_attrs:
                continue
            normal_attrs.append(f'{name}="{escape(value, {"\"": "&quot;"})}"')
        if not any(attr.startswith("type=") for attr in normal_attrs):
            normal_attrs.insert(0, 'type="list"')
        normal_attrs.append(f'sqref="{escape(sqref_match.group(1), {"\"": "&quot;"})}"')
        formula = escape(formula_match.group(1))
        fragments.append(f"<dataValidation {' '.join(normal_attrs)}><formula1>{formula}</formula1></dataValidation>")
    return fragments


def _data_validation_signature(fragment: str) -> tuple[str, str] | None:
    sqref_match = re.search(r"\bsqref=\"([^\"]+)\"", fragment)
    formula_match = re.search(r"<formula1>(.*?)</formula1>", fragment, flags=re.DOTALL)
    if not sqref_match or not formula_match:
        return None
    return sqref_match.group(1), formula_match.group(1)


def _inject_normal_data_validations(xml: str, fragments: list[str]) -> str:
    if not fragments:
        return xml

    existing = {
        signature
        for signature in (_data_validation_signature(match.group(0)) for match in re.finditer(r"<dataValidation\b.*?</dataValidation>", xml, flags=re.DOTALL))
        if signature is not None
    }
    additions = [fragment for fragment in fragments if _data_validation_signature(fragment) not in existing]
    if not additions:
        return xml

    section_match = re.search(r"<dataValidations\b[^>]*>.*?</dataValidations>", xml, flags=re.DOTALL)
    if section_match:
        section = section_match.group(0)
        updated_section = section.replace("</dataValidations>", "".join(additions) + "</dataValidations>", 1)
        count = len(re.findall(r"<dataValidation\b", updated_section))
        if re.search(r"\bcount=\"\d+\"", updated_section):
            updated_section = re.sub(r"\bcount=\"\d+\"", f'count="{count}"', updated_section, count=1)
        else:
            updated_section = updated_section.replace("<dataValidations", f'<dataValidations count="{count}"', 1)
        return xml[: section_match.start()] + updated_section + xml[section_match.end() :]

    section = f'<dataValidations count="{len(additions)}">{"".join(additions)}</dataValidations>'
    if "<extLst" in xml:
        return xml.replace("<extLst", section + "<extLst", 1)
    return xml.replace("</worksheet>", section + "</worksheet>", 1)


def _restore_sheet_extensions(source_path: Path, target_path: Path, sheet_name: str) -> None:
    sheet_path = _sheet_xml_path(source_path, sheet_name)
    if sheet_path is None:
        return

    with zipfile.ZipFile(source_path) as source_archive:
        try:
            source_xml = source_archive.read(sheet_path).decode("utf-8")
        except KeyError:
            return
    source_ext = _worksheet_ext_list(source_xml)
    source_opening_tag = _worksheet_opening_tag(source_xml)
    validation_fragments = _x14_data_validation_fragments(source_xml)
    if not source_ext or "dataValidations" not in source_ext or not source_opening_tag:
        return

    with zipfile.ZipFile(target_path) as target_archive:
        try:
            target_xml = target_archive.read(sheet_path).decode("utf-8")
        except KeyError:
            return
        entries = [(info, target_archive.read(info.filename)) for info in target_archive.infolist()]

    target_opening_tag = _worksheet_opening_tag(target_xml)
    if target_opening_tag:
        target_xml = target_xml.replace(target_opening_tag, source_opening_tag, 1)
    if _worksheet_ext_list(target_xml):
        target_xml = re.sub(r"<extLst\b.*?</extLst>", source_ext, target_xml, count=1, flags=re.DOTALL)
    else:
        target_xml = target_xml.replace("</worksheet>", f"{source_ext}</worksheet>", 1)
    target_xml = _inject_normal_data_validations(target_xml, validation_fragments)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = Path(tmp.name)
    try:
        with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as output_archive:
            for info, data in entries:
                if info.filename == sheet_path:
                    output_archive.writestr(info, target_xml.encode("utf-8"))
                else:
                    output_archive.writestr(info, data)
        shutil.move(tmp_path, target_path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def _copy_cell(source, target, formula_template: tuple[object, int] | None = None) -> None:
    value = source.value
    if _is_formula(value):
        value = _translate_formula(value, source, target)
    elif formula_template is not None and value in (None, ""):
        formula_value, template_row = formula_template
        formula_source = source.parent.cell(row=template_row, column=source.column)
        value = _translate_formula(formula_value, formula_source, target)
    target.value = value
    if source.has_style:
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)
    if source.hyperlink:
        target._hyperlink = copy.copy(source.hyperlink)
    if source.comment:
        target.comment = copy.copy(source.comment)


def _write_group_workbook(source_path: Path, target_path: Path, sheet_name: str, field_column: int, group_value: str) -> int:
    source_workbook = load_workbook(source_path, data_only=False)
    target_workbook = load_workbook(target_path, data_only=False)
    try:
        source_sheet = source_workbook[sheet_name]
        target_sheet = target_workbook[sheet_name]
        data_columns = [
            cell.column
            for cell in source_sheet[HEADER_ROW]
            if _clean_header(cell.value) in ALLOWED_SPLIT_FIELDS
        ]
        formula_templates = _detect_formula_templates(source_sheet)
        kept = 0
        for source_row in _iter_person_data_rows(source_sheet):
            row_values = [source_sheet.cell(row=source_row, column=column).value for column in data_columns]
            if not any(value not in (None, "") for value in row_values):
                continue
            value = str(source_sheet.cell(row=source_row, column=field_column).value or "").strip() or "未填写"
            if value != group_value:
                continue
            target_row = DATA_START_ROW + kept
            for column in range(1, source_sheet.max_column + 1):
                _copy_cell(
                    source_sheet.cell(row=source_row, column=column),
                    target_sheet.cell(row=target_row, column=column),
                    formula_templates.get(column),
                )
            if source_row in source_sheet.row_dimensions:
                target_sheet.row_dimensions[target_row].height = source_sheet.row_dimensions[source_row].height
            kept += 1

        first_extra_row = DATA_START_ROW + kept
        if first_extra_row <= target_sheet.max_row:
            target_sheet.delete_rows(first_extra_row, target_sheet.max_row - first_extra_row + 1)

        target_workbook.save(target_path)
        _restore_sheet_extensions(source_path, target_path, sheet_name)
        return kept
    finally:
        source_workbook.close()
        target_workbook.close()


def split_workbook_by_field(workbook_path: Path, output_dir: Path, field_name: str, sheet_name: str | None = None) -> SplitResult:
    if field_name not in ALLOWED_SPLIT_FIELDS:
        raise ValueError("请选择系统支持的拆分字段。")
    available_sheets = list_split_sheets(workbook_path)
    if sheet_name is None:
        sheet_name = available_sheets[0].name
    matched_sheet = next((sheet for sheet in available_sheets if sheet.name == sheet_name), None)
    if matched_sheet is None:
        raise ValueError(f"未找到可用于拆分的 Sheet「{sheet_name}」。")
    available_fields = {field.name for field in matched_sheet.fields}
    if field_name not in available_fields:
        raise ValueError(f"Sheet「{sheet_name}」第 3 行不包含拆分字段「{field_name}」。")

    field_column, groups = _group_values(workbook_path, sheet_name, field_name)
    split_dir = output_dir / "split_excels"
    split_dir.mkdir(parents=True, exist_ok=True)

    generated: list[dict[str, object]] = []
    for index, group_value in enumerate(sorted(groups.keys()), start=1):
        filename = f"{index:02d}_{_safe_filename_part(field_name)}_{_safe_filename_part(group_value)}.xlsx"
        target_path = split_dir / filename
        shutil.copy2(workbook_path, target_path)
        kept = _write_group_workbook(workbook_path, target_path, sheet_name, field_column, group_value)
        generated.append({"字段值": group_value, "人数": kept, "文件名": filename})

    output_file = output_dir / f"人才盘点拆分结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    with zipfile.ZipFile(output_file, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for item in generated:
            archive.write(split_dir / str(item["文件名"]), arcname=str(item["文件名"]))

    summary = {
        "处理类型": "表格拆分",
        "拆分依据": field_name,
        "拆分 Sheet": sheet_name,
        "生成文件数": len(generated),
        "总人员数": sum(int(item["人数"]) for item in generated),
        "拆分明细": generated,
    }
    return SplitResult(output_file=output_file, summary=summary)
