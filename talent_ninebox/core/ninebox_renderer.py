from __future__ import annotations

from math import ceil

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import RowRecord
from .ninebox_mapper import NINEBOX_KEYS

FILLS = {
    "高潜力 / 高绩效": "C6E0B4",
    "高潜力 / 中绩效": "E2F0D9",
    "高潜力 / 低绩效": "FFF2CC",
    "中潜力 / 高绩效": "D9EAD3",
    "中潜力 / 中绩效": "DDEBF7",
    "中潜力 / 低绩效": "FCE4D6",
    "低潜力 / 高绩效": "E2F0D9",
    "低潜力 / 中绩效": "FCE4D6",
    "低潜力 / 低绩效": "F4CCCC",
}


VALUE_HIGH_FILL = PatternFill("solid", fgColor="00B050")
VALUE_LOW_FILL = PatternFill("solid", fgColor="FF0000")
HIGH_RISK_FONT = Font(color="9C0006", bold=True)
DEFAULT_NAME_FONT = Font(color="1F2933")


def _display_box_label(key: str) -> str:
    return key.replace("潜力", "能力")


def _format_ratio(count: int, total: int) -> str:
    if total <= 0:
        return "0%"
    ratio = count / total
    text = f"{ratio:.1%}"
    return text.replace(".0%", "%")


def render_ninebox(
    ws: Worksheet,
    people_by_box: dict[str, list[RowRecord]],
    placement_field: str = "",
    total_people: int | None = None,
) -> None:
    people_columns_per_box = 4
    denominator = total_people if total_people is not None else sum(len(people) for people in people_by_box.values())
    max_people = max((len(people) for people in people_by_box.values()), default=0)
    people_rows_per_box = max(6, ceil(max_people / people_columns_per_box))
    box_height = people_rows_per_box + 1
    row_gap = 1
    row_starts = {
        "高潜力": 3,
        "中潜力": 3 + box_height + row_gap,
        "低潜力": 3 + (box_height + row_gap) * 2,
    }
    col_starts = {
        "低绩效": 2,
        "中绩效": 6,
        "高绩效": 10,
    }
    grid_positions = {
        "高潜力 / 低绩效": (row_starts["高潜力"], col_starts["低绩效"]),
        "高潜力 / 中绩效": (row_starts["高潜力"], col_starts["中绩效"]),
        "高潜力 / 高绩效": (row_starts["高潜力"], col_starts["高绩效"]),
        "中潜力 / 低绩效": (row_starts["中潜力"], col_starts["低绩效"]),
        "中潜力 / 中绩效": (row_starts["中潜力"], col_starts["中绩效"]),
        "中潜力 / 高绩效": (row_starts["中潜力"], col_starts["高绩效"]),
        "低潜力 / 低绩效": (row_starts["低潜力"], col_starts["低绩效"]),
        "低潜力 / 中绩效": (row_starts["低潜力"], col_starts["中绩效"]),
        "低潜力 / 高绩效": (row_starts["低潜力"], col_starts["高绩效"]),
    }

    ws.title = "人才九宫格"
    ws.sheet_view.showGridLines = False
    ws["B1"] = f"人才九宫格（依据：{placement_field or '九宫格落位'}）"
    ws["B1"].font = Font(size=18, bold=True)
    ws["B2"] = "绩效：低 → 中 → 高"
    ws["A3"] = "能力：高"
    ws.cell(row_starts["中潜力"], 1).value = "能力：中"
    ws.cell(row_starts["低潜力"], 1).value = "能力：低"
    for row in row_starts.values():
        ws.cell(row, 1).alignment = Alignment(text_rotation=90, horizontal="center", vertical="center")

    legend_row = row_starts["低潜力"] + box_height + 1
    ws.cell(legend_row, 2).value = "标记说明：离职风险高=姓名红字；价值观综合得分高=姓名单元格绿色；价值观综合得分低=姓名单元格红色。"
    ws.cell(legend_row, 2).font = Font(color="666666")

    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 18
    for row in range(3, legend_row + 1):
        ws.row_dimensions[row].height = 34

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for key in NINEBOX_KEYS:
        start_row, start_col = grid_positions[key]
        people = people_by_box.get(key, [])
        fill = PatternFill("solid", fgColor=FILLS[key])
        end_col = start_col + 3
        end_row = start_row + people_rows_per_box

        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
        title = ws.cell(start_row, start_col)
        title.value = f"{_display_box_label(key)}（{len(people)}人，{_format_ratio(len(people), denominator)}）"
        title.font = Font(bold=True)
        title.fill = fill
        title.alignment = Alignment(horizontal="center", vertical="center")

        idx = 0
        for row in range(start_row + 1, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row, col)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if idx < len(people):
                    person = people[idx]
                    cell.value = person.display_name
                    if person.value_score_level == "high":
                        cell.fill = VALUE_HIGH_FILL
                    elif person.value_score_level == "low":
                        cell.fill = VALUE_LOW_FILL
                    cell.font = HIGH_RISK_FONT if person.high_attrition_risk else DEFAULT_NAME_FONT
                idx += 1

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row, col).border = border

    ws.freeze_panes = "B3"
