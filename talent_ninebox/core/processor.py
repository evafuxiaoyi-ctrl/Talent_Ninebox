from __future__ import annotations

import copy
import re
import tempfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.worksheet.formula import ArrayFormula

from .field_utils import find_field_index, find_placement_index, normalize_field, placement_field_label
from .input_handler import extract_excels
from .models import Issue, ProcessOptions, ProcessResult, ProcessSummary, RowRecord, WorkbookInfo
from .ninebox_mapper import NINEBOX_KEYS, map_ninebox
from .ninebox_renderer import render_ninebox
from .report_writer import write_instructions, write_issues, write_summary, write_table
from .workbook_inspector import inspect_workbook

SOURCE_COLUMNS = ["来源文件", "来源Sheet", "来源行号", "是否重复", "数据异常提示"]
GENERATED_SHEETS = {"使用说明", "整合总表", "人才九宫格", "处理摘要", "异常报告", "文件来源映射"}
STRONG_GREEN_FILL = PatternFill("solid", fgColor="00B050")
STRONG_RED_FILL = PatternFill("solid", fgColor="FF0000")
WHITE_FONT = Font(color="FFFFFF", bold=True)
SAMPLE_ROW_MARKERS = {"示例行"}
RED_FONT_COLORS = {"FFC00000", "FFFF0000", "00FF0000"}


def _index(headers: list[str], name: str) -> int | None:
    return find_field_index(headers, name)


def _placement_index(headers: list[str], options: ProcessOptions) -> int | None:
    return find_placement_index(headers, options.placement_mode)


def _numeric(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def _value_score(values: list[Any], headers: list[str]) -> float | None:
    score_idx = _index(headers, "价值观综合得分")
    score = _numeric(values[score_idx]) if score_idx is not None and score_idx < len(values) else None
    if score is None:
        component_scores: list[float] = []
        for field in ["成就客户", "激情人生", "拥抱变化", "团队协作"]:
            idx = _index(headers, field)
            if idx is None or idx >= len(values):
                continue
            component = _numeric(values[idx])
            if component is None:
                component = _leading_score(values[idx])
            if component is not None:
                component_scores.append(component)
        if component_scores:
            score = sum(component_scores) / len(component_scores)
    return score


def _value_score_level(values: list[Any], headers: list[str]) -> str:
    score = _value_score(values, headers)
    if score is None:
        return ""
    if score >= 4:
        return "high"
    if score <= 2:
        return "low"
    return "middle"


def _is_high_attrition_risk(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return text in {"高", "高风险", "高离职风险", "High", "high"}


def _display_name(values: list[Any], headers: list[str]) -> str:
    name_idx = _index(headers, "姓名")
    first_dept_idx = _index(headers, "一级部门")
    second_dept_idx = _index(headers, "二级部门")
    name = values[name_idx] if name_idx is not None and name_idx < len(values) else ""
    if name in (None, ""):
        return ""
    parts = [str(name).strip()]
    for idx in [first_dept_idx, second_dept_idx]:
        if idx is None or idx >= len(values):
            continue
        value = values[idx]
        if value not in (None, ""):
            parts.append(str(value).strip())
    return "-".join(parts)


def _leading_score(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.match(r"\s*([1-5])", str(value))
    if not match:
        return None
    return float(match.group(1))


def _avg(values: list[float]) -> float | None:
    return sum(values) / len(values) if values else None


def _level_from_score(score: float | None) -> str | None:
    if score is None:
        return None
    if score >= 4:
        return "高"
    if score >= 3:
        return "中"
    return "低"


def _performance_level(values: list[Any], headers: list[str]) -> str | None:
    for field in ["绩效九宫格落位"]:
        idx = _index(headers, field)
        if idx is not None and idx < len(values) and values[idx] in {"高", "中", "低"}:
            return str(values[idx])
    for field in ["25+26绩效综合等级", "25年绩效等级"]:
        idx = _index(headers, field)
        if idx is None or idx >= len(values):
            continue
        value = str(values[idx]).strip()
        if value in {"A", "B+"}:
            return "高"
        if value == "B":
            return "中"
        if value in {"C", "D"}:
            return "低"
    return None


def _ability_level(values: list[Any], headers: list[str]) -> str | None:
    ability_idx = _index(headers, "能力九宫格落位")
    if ability_idx is not None and ability_idx < len(values) and values[ability_idx] in {"高", "中", "低"}:
        return str(values[ability_idx])

    score_idx = _index(headers, "能力综合评分")
    score = _numeric(values[score_idx]) if score_idx is not None and score_idx < len(values) else None
    if score is not None:
        return _level_from_score(score)

    leadership_fields = ["【原一：经营思维】", "【原二：战略能力】", "【原三：用户思维】", "【原四：系统思考】", "【原五：体系机制】", "【原六：超越伯乐】"]
    professional_fields = ["专业知识", "专业洞察", "专业影响"]
    leadership_scores = []
    professional_scores = []
    for field in leadership_fields:
        idx = _index(headers, field)
        if idx is not None and idx < len(values):
            score = _leading_score(values[idx])
            if score is not None:
                leadership_scores.append(score)
    for field in professional_fields:
        idx = _index(headers, field)
        if idx is not None and idx < len(values):
            score = _leading_score(values[idx])
            if score is not None:
                professional_scores.append(score)

    component_scores = [score for score in [_avg(leadership_scores), _avg(professional_scores)] if score is not None]
    return _level_from_score(_avg(component_scores))


def _derived_initial_placement(values: list[Any], headers: list[str]) -> str | None:
    performance = _performance_level(values, headers)
    ability = _ability_level(values, headers)
    if not performance or not ability:
        return None
    return f"{performance}绩效{ability}能力"


def _calculated_values(record: RowRecord) -> list[Any]:
    if not record.calculated_values:
        return record.values
    return [
        calculated if calculated not in (None, "") else record.values[idx]
        for idx, calculated in enumerate(record.calculated_values)
    ]


def _placement_value(record: RowRecord, headers: list[str], options: ProcessOptions, place_idx: int) -> Any:
    values = _calculated_values(record)
    raw_value = values[place_idx] if place_idx < len(values) else None
    if map_ninebox(raw_value):
        return raw_value
    if options.placement_mode == "initial":
        return _derived_initial_placement(values, headers) or _derived_initial_placement(record.values, headers) or raw_value
    return raw_value


def _copy_cell_style(source, target) -> None:
    if source.has_style:
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)
    else:
        target._style = copy.copy(source._style)


def _is_sample_row(ws, row_idx: int) -> bool:
    return str(ws.cell(row_idx, 1).value or "").strip() in SAMPLE_ROW_MARKERS


def _font_color_key(cell) -> str:
    color = cell.font.color
    if color is None or color.type != "rgb" or color.rgb is None:
        return ""
    return str(color.rgb).upper()


def _row_looks_like_red_sample_style(ws, row_idx: int, max_col: int) -> bool:
    red_cells = sum(1 for col in range(1, max_col + 1) if _font_color_key(ws.cell(row_idx, col)) in RED_FONT_COLORS)
    return red_cells > max_col / 2


def _choose_data_style_row(ws, header_row: int, max_col: int) -> int:
    fallback = min(header_row + 1, ws.max_row)
    first_data_row = None
    for row_idx in range(header_row + 1, min(ws.max_row, header_row + 10) + 1):
        if _is_sample_row(ws, row_idx):
            continue
        first_data_row = first_data_row or row_idx
        if not _row_looks_like_red_sample_style(ws, row_idx, max_col):
            return row_idx
    return first_data_row or fallback


def _copy_worksheet(source_ws, target_wb: Workbook, title: str | None = None):
    target_ws = target_wb.create_sheet(title or source_ws.title)
    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws.cell(source_cell.row, source_cell.column, source_cell.value)
            _copy_cell_style(source_cell, target_cell)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy.copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy.copy(source_cell.comment)

    for col_letter, dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dimension.width
        target_ws.column_dimensions[col_letter].hidden = dimension.hidden

    for row_idx, dimension in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = dimension.height
        target_ws.row_dimensions[row_idx].hidden = dimension.hidden

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    return target_ws


def _prepare_output_workbook(template_info: WorkbookInfo) -> Workbook:
    wb = load_workbook(template_info.path, data_only=False)
    for sheet_name in list(wb.sheetnames):
        if sheet_name in GENERATED_SHEETS and sheet_name != template_info.sheet_name:
            wb.remove(wb[sheet_name])
    return wb


def _copy_support_sheets(wb: Workbook, template_info: WorkbookInfo) -> None:
    template_wb = load_workbook(template_info.path, data_only=False)
    for source_ws in template_wb.worksheets:
        if source_ws.title == template_info.sheet_name:
            continue
        if source_ws.title in GENERATED_SHEETS:
            continue
        title = source_ws.title
        if title in wb.sheetnames:
            title = f"原模板_{title}"
        _copy_worksheet(source_ws, wb, title)


def _copy_data_validations(template_ws, target_ws, template_header_row: int, record_count: int, max_col: int) -> None:
    if record_count <= 0:
        return
    target_start_row = template_header_row + 1
    target_end_row = template_header_row + record_count
    template_data_start = template_header_row + 1

    for source_validation in template_ws.data_validations.dataValidation:
        columns: set[int] = set()
        for cell_range in source_validation.ranges.ranges:
            if cell_range.max_row < template_data_start:
                continue
            for col in range(cell_range.min_col, min(cell_range.max_col, max_col) + 1):
                columns.add(col)
        if not columns:
            continue

        target_validation = copy.copy(source_validation)
        target_validation.ranges = MultiCellRange()
        for col in sorted(columns):
            letter = get_column_letter(col)
            target_validation.add(f"{letter}{target_start_row}:{letter}{target_end_row}")
        target_ws.add_data_validation(target_validation)


def _choose_template(infos: list[WorkbookInfo]) -> tuple[tuple[str, ...], int]:
    counts = Counter(info.signature for info in infos)
    return counts.most_common(1)[0]


def _read_records(info: WorkbookInfo, headers: list[str], issues: list[Issue], options: ProcessOptions) -> list[RowRecord]:
    wb = load_workbook(info.path, data_only=False)
    value_wb = load_workbook(info.path, data_only=True)
    ws = wb[info.sheet_name]
    value_ws = value_wb[info.sheet_name]
    try:
        records: list[RowRecord] = []
        name_idx = _index(headers, "姓名")
        id_idx = _index(headers, "工号")
        place_idx = _placement_index(headers, options)
        risk_idx = _index(headers, "离职风险")
        max_col = len(headers)

        for row_idx in range(info.header_row + 1, ws.max_row + 1):
            if _is_sample_row(ws, row_idx):
                continue
            values = [ws.cell(row_idx, col).value for col in range(1, max_col + 1)]
            calculated_values = [value_ws.cell(row_idx, col).value for col in range(1, max_col + 1)]
            effective_values = [
                calculated if calculated not in (None, "") else values[idx]
                for idx, calculated in enumerate(calculated_values)
            ]
            has_name = name_idx is not None and name_idx < len(effective_values) and effective_values[name_idx] not in (None, "")
            has_id = id_idx is not None and id_idx < len(effective_values) and effective_values[id_idx] not in (None, "")
            if not has_name and not has_id:
                continue
            if name_idx is not None and effective_values[name_idx] in (None, ""):
                issues.append(Issue("warning", "数据异常", "姓名为空", info.source_name, info.sheet_name, row_idx, "姓名", effective_values[name_idx]))
            if place_idx is not None and effective_values[place_idx] in (None, ""):
                issues.append(Issue("warning", "九宫格落位异常", "人才九宫格落位为空", info.source_name, info.sheet_name, row_idx, "人才九宫格落位", effective_values[place_idx]))
            if id_idx is not None or name_idx is not None:
                record = RowRecord(values, info.source_name, info.sheet_name, row_idx, calculated_values=calculated_values)
                record.display_name = _display_name(effective_values, headers)
                if risk_idx is not None and risk_idx < len(effective_values):
                    record.high_attrition_risk = _is_high_attrition_risk(effective_values[risk_idx])
                record.value_score_level = _value_score_level(effective_values, headers)
                records.append(record)
        return records
    finally:
        wb.close()
        value_wb.close()


def _mark_duplicates(records: list[RowRecord], headers: list[str], issues: list[Issue]) -> None:
    id_idx = _index(headers, "工号")
    name_idx = _index(headers, "姓名")
    dept_idx = _index(headers, "所属部门")
    seen: dict[tuple[Any, ...], RowRecord] = {}
    for record in records:
        key: tuple[Any, ...] | None = None
        if id_idx is not None and record.values[id_idx] not in (None, ""):
            key = ("id", str(record.values[id_idx]).strip())
        elif name_idx is not None and record.values[name_idx] not in (None, ""):
            dept = record.values[dept_idx] if dept_idx is not None else ""
            key = ("name_dept", str(record.values[name_idx]).strip(), str(dept or "").strip())
        if key is None:
            continue
        if key in seen:
            record.is_duplicate = True
            record.notes.append("疑似重复")
            issues.append(Issue("warning", "数据异常", "疑似重复人员", record.source_file, record.sheet, record.source_row, "工号/姓名", key))
        else:
            seen[key] = record


def _is_formula(value: object) -> bool:
    return (isinstance(value, str) and value.startswith("=")) or isinstance(value, ArrayFormula)


def _translate_formula(value: object, source, target) -> object:
    if isinstance(value, ArrayFormula):
        origin = value.ref or source.coordinate
        text = value.text
    else:
        origin = source.coordinate
        text = value
    if not isinstance(text, str):
        return value
    if source.coordinate != target.coordinate:
        try:
            text = Translator(text, origin=origin).translate_formula(target.coordinate)
        except Exception:
            pass
    if isinstance(value, ArrayFormula):
        return ArrayFormula(target.coordinate, text)
    return text


def _detect_formula_templates(template_info: WorkbookInfo, headers: list[str]) -> dict[int, tuple[object, int]]:
    wb = load_workbook(template_info.path, data_only=False)
    ws = wb[template_info.sheet_name]
    formulas: dict[int, tuple[object, int]] = {}
    for col in range(1, len(headers) + 1):
        for row in range(template_info.header_row + 1, min(ws.max_row, template_info.header_row + 10) + 1):
            value = ws.cell(row, col).value
            if _is_formula(value):
                formulas[col] = (copy.copy(value), row)
                break
    return formulas


def _write_merged_sheet(wb: Workbook, template_info: WorkbookInfo, records: list[RowRecord], options: ProcessOptions) -> None:
    template_wb = load_workbook(template_info.path, data_only=False)
    template_ws = template_wb[template_info.sheet_name]
    if template_info.sheet_name in wb.sheetnames:
        ws = wb[template_info.sheet_name]
        copied_from_template = True
    else:
        ws = wb.create_sheet(template_info.sheet_name)
        copied_from_template = False
    headers = template_info.headers[:]
    output_headers = headers + (SOURCE_COLUMNS if options.add_source_columns else [])
    header_row = template_info.header_row
    output_start_row = header_row + 1

    for row in range(1, header_row):
        for col in range(1, len(headers) + 1):
            source = template_ws.cell(row, col)
            target = ws.cell(row, col, source.value)
            _copy_cell_style(source, target)
        if template_ws.row_dimensions[row].height:
            ws.row_dimensions[row].height = template_ws.row_dimensions[row].height

    for merged_range in template_ws.merged_cells.ranges:
        if merged_range.max_row < header_row:
            ws.merge_cells(str(merged_range))

    for idx, header in enumerate(output_headers, start=1):
        cell = ws.cell(header_row, idx, header)
        source_cell = template_ws.cell(template_info.header_row, idx if idx <= len(headers) else len(headers))
        _copy_cell_style(source_cell, cell)

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = template_ws.column_dimensions[letter].width or 12

    formula_templates = _detect_formula_templates(template_info, headers)
    data_style_row = _choose_data_style_row(template_ws, template_info.header_row, len(headers))
    value_score_idx = _index(headers, "价值观综合得分")
    risk_idx = _index(headers, "离职风险")

    for out_row, record in enumerate(records, start=output_start_row):
        value_score = _value_score(record.values, headers)
        for col in range(1, len(headers) + 1):
            target = ws.cell(out_row, col)
            style_source = template_ws.cell(data_style_row, col)
            _copy_cell_style(style_source, target)
            if headers[col - 1] == "序号":
                target.value = out_row - header_row
            elif col in formula_templates:
                formula, source_row = formula_templates[col]
                source = template_ws.cell(source_row, col)
                target.value = _translate_formula(formula, source, target)
            else:
                target.value = record.values[col - 1] if col - 1 < len(record.values) else None

            idx = col - 1
            if value_score_idx == idx and value_score is not None:
                if value_score >= 4:
                    target.fill = STRONG_GREEN_FILL
                    target.font = WHITE_FONT
                elif value_score <= 2:
                    target.fill = STRONG_RED_FILL
                    target.font = WHITE_FONT
            if risk_idx == idx and record.high_attrition_risk:
                target.fill = STRONG_RED_FILL
                target.font = WHITE_FONT

        if options.add_source_columns:
            extras = [record.source_file, record.sheet, record.source_row, "是" if record.is_duplicate else "否", "；".join(record.notes)]
            for offset, value in enumerate(extras, start=len(headers) + 1):
                ws.cell(out_row, offset, value)

    if not copied_from_template:
        _copy_data_validations(template_ws, ws, template_info.header_row, len(records), len(headers))
    first_extra_row = output_start_row + len(records)
    if first_extra_row <= ws.max_row:
        ws.delete_rows(first_extra_row, ws.max_row - first_extra_row + 1)
    ws.freeze_panes = f"A{output_start_row}"
    ws.auto_filter.ref = ws.dimensions


def _build_ninebox(records: list[RowRecord], headers: list[str], issues: list[Issue], options: ProcessOptions) -> dict[str, list[RowRecord]]:
    name_idx = _index(headers, "姓名")
    place_idx = _placement_index(headers, options)
    people_by_box: dict[str, list[RowRecord]] = {key: [] for key in NINEBOX_KEYS}
    if name_idx is None or place_idx is None:
        return people_by_box
    for record in records:
        name = record.values[name_idx]
        place = _placement_value(record, headers, options, place_idx)
        if name in (None, ""):
            continue
        box = map_ninebox(place)
        if box:
            record.ninebox_key = box
            if not record.display_name:
                record.display_name = str(name)
            people_by_box[box].append(record)
        else:
            record.notes.append("九宫格落位未识别")
            issues.append(Issue("warning", "九宫格落位异常", "九宫格落位无法识别", record.source_file, record.sheet, record.source_row, placement_field_label(headers, options.placement_mode), place))
    return people_by_box


def _process_infos(
    infos: list[WorkbookInfo],
    issues: list[Issue],
    output_dir: Path,
    summary: ProcessSummary,
    options: ProcessOptions,
    output_prefix: str,
) -> ProcessResult:
    if not infos:
        detail = ""
        if issues:
            first = issues[0]
            location = f"文件「{first.source_file}」" if first.source_file else "输入文件"
            detail = f"{location}：{first.message}"
        raise ValueError(f"没有可处理的 Excel 文件。{detail}")

    main_signature, _ = _choose_template(infos)
    template_info = next(info for info in infos if info.signature == main_signature)
    headers = template_info.headers
    placement_field = placement_field_label(headers, options.placement_mode)
    summary.stage_label = options.stage_label
    summary.placement_field = placement_field

    records: list[RowRecord] = []
    processed_sources = set()
    for info in infos:
        if info.signature != main_signature:
            issues.append(Issue("error", "模板异常", "表头结构与主模板不一致，已跳过", info.source_name, info.sheet_name, info.header_row))
            continue
        records.extend(_read_records(info, headers, issues, options))
        processed_sources.add(info.source_name)

    summary.processed_file_count = len(processed_sources)
    summary.skipped_file_count = summary.excel_file_count - summary.processed_file_count
    if summary.processed_file_count == 0:
        first = next((issue for issue in issues if issue.category == "模板异常"), issues[0] if issues else None)
        detail = ""
        if first:
            location = f"文件「{first.source_file}」"
            if first.sheet:
                location += f" / Sheet「{first.sheet}」"
            if first.row:
                location += f" / 第 {first.row} 行"
            detail = f"{location}：{first.message}"
        raise ValueError(f"所有 Excel 都因模板不一致被跳过，未生成结果。{detail}")

    _mark_duplicates(records, headers, issues)
    people_by_box = _build_ninebox(records, headers, issues, options)

    wb = _prepare_output_workbook(template_info)
    if "使用说明" in wb.sheetnames:
        wb.remove(wb["使用说明"])
    write_instructions(wb.create_sheet("使用说明", 0))
    _write_merged_sheet(wb, template_info, records, options)
    render_ninebox(wb.create_sheet("人才九宫格"), people_by_box, placement_field, total_people=len(records))

    summary.total_people = len(records)
    summary.placed_people = sum(len(v) for v in people_by_box.values())
    summary.unplaced_people = max(summary.total_people - summary.placed_people, 0)
    summary.duplicate_count = sum(1 for r in records if r.is_duplicate)
    summary.issue_count = len(issues)
    summary.ninebox_counts = {key: len(people_by_box.get(key, [])) for key in NINEBOX_KEYS}

    source_rows = []
    name_idx = _index(headers, "姓名")
    id_idx = _index(headers, "工号")
    for idx, record in enumerate(records, start=template_info.header_row + 1):
        source_rows.append([
            idx,
            record.values[name_idx] if name_idx is not None else "",
            record.values[id_idx] if id_idx is not None else "",
            record.source_file,
            record.sheet,
            record.source_row,
        ])

    filename = f"{output_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_file = output_dir / filename
    summary.output_file = filename
    write_summary(wb.create_sheet("处理摘要"), summary)
    write_issues(wb.create_sheet("异常报告"), issues)
    write_table(wb.create_sheet("文件来源映射"), ["合并后行号", "姓名", "工号", "来源文件", "来源Sheet", "原始行号"], source_rows)
    wb.save(output_file)
    return ProcessResult(output_file, summary, issues)


def process_zip(zip_path: str | Path, output_dir: str | Path, options: ProcessOptions | None = None) -> ProcessResult:
    options = options or ProcessOptions()
    zip_path = Path(zip_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    issues: list[Issue] = []
    summary = ProcessSummary(upload_file=zip_path.name)

    with tempfile.TemporaryDirectory(prefix="talent-ninebox-") as tmp:
        work_dir = Path(tmp)
        extracted, extract_issues = extract_excels(zip_path, work_dir, options)
        issues.extend(extract_issues)
        summary.excel_file_count = len(extracted)

        infos: list[WorkbookInfo] = []
        for source_name, path in extracted:
            info, inspect_issues = inspect_workbook(source_name, path, options)
            issues.extend(inspect_issues)
            if info:
                infos.append(info)

        return _process_infos(infos, issues, output_dir, summary, options, "人才盘点整合结果_初版")


def process_workbook(workbook_path: str | Path, output_dir: str | Path, options: ProcessOptions | None = None) -> ProcessResult:
    options = options or ProcessOptions(placement_mode="final", stage_label="终版九宫格生成", add_source_columns=False)
    workbook_path = Path(workbook_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("终版生成仅支持上传 .xlsx 文件。")

    issues: list[Issue] = []
    summary = ProcessSummary(upload_file=workbook_path.name, excel_file_count=1)
    info, inspect_issues = inspect_workbook(workbook_path.name, workbook_path, options)
    issues.extend(inspect_issues)
    infos = [info] if info else []
    return _process_infos(infos, issues, output_dir, summary, options, "人才盘点整合结果_终版")
