from __future__ import annotations

import copy
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import GenericMergeResult, GenericTemplateConfig
from .splitter import _copy_cell, _row_has_data, _safe_filename_part

SOURCE_COLUMNS = ["来源文件", "来源Sheet", "来源行号"]


def _copy_header_extras(sheet, source_start_col: int) -> None:
    for offset, header in enumerate(SOURCE_COLUMNS, start=source_start_col):
        cell = sheet.cell(row=1, column=offset, value=header)
        source = sheet.cell(row=1, column=max(1, source_start_col - 1))
        if source.has_style:
            cell.font = copy.copy(source.font)
            cell.fill = copy.copy(source.fill)
            cell.border = copy.copy(source.border)
            cell.alignment = copy.copy(source.alignment)
            cell.number_format = source.number_format


def _clear_data_rows(sheet, data_start_row: int) -> None:
    if data_start_row <= sheet.max_row:
        sheet.delete_rows(data_start_row, sheet.max_row - data_start_row + 1)


def merge_workbooks_by_config(
    workbook_paths: list[str | Path],
    output_dir: str | Path,
    config: GenericTemplateConfig,
    append_source_columns: bool = True,
) -> GenericMergeResult:
    paths = [Path(path) for path in workbook_paths]
    if not paths:
        raise ValueError("请至少提供一个待合并的 Excel 文件。")

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / f"{_safe_filename_part(config.name)}合并结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy2(paths[0], output_file)

    target_workbook = load_workbook(output_file, data_only=False)
    try:
        sheet_row_counts: dict[str, int] = {}
        source_rows: list[list[Any]] = []
        for sheet_config in config.sheets:
            if sheet_config.name not in target_workbook.sheetnames:
                raise ValueError(f"模板文件缺少 Sheet「{sheet_config.name}」。")
            target_sheet = target_workbook[sheet_config.name]
            base_max_col = target_sheet.max_column
            if append_source_columns:
                _copy_header_extras(target_sheet, base_max_col + 1)
            _clear_data_rows(target_sheet, sheet_config.data_start_row)

            target_row = sheet_config.data_start_row
            for path in paths:
                source_workbook = load_workbook(path, data_only=False)
                try:
                    if sheet_config.name not in source_workbook.sheetnames:
                        raise ValueError(f"文件「{path.name}」缺少 Sheet「{sheet_config.name}」。")
                    source_sheet = source_workbook[sheet_config.name]
                    for source_row in range(sheet_config.data_start_row, source_sheet.max_row + 1):
                        if not _row_has_data(source_sheet, source_row, source_sheet.max_column):
                            continue
                        for column in range(1, base_max_col + 1):
                            _copy_cell(source_sheet.cell(row=source_row, column=column), target_sheet.cell(row=target_row, column=column))
                        if append_source_columns:
                            extras = [path.name, sheet_config.name, source_row]
                            for offset, value in enumerate(extras, start=base_max_col + 1):
                                target_sheet.cell(row=target_row, column=offset, value=value)
                        source_rows.append([sheet_config.name, target_row, path.name, source_row])
                        target_row += 1
                finally:
                    source_workbook.close()

            sheet_row_counts[sheet_config.name] = target_row - sheet_config.data_start_row

        if "处理摘要" in target_workbook.sheetnames:
            del target_workbook["处理摘要"]
        summary_sheet = target_workbook.create_sheet("处理摘要")
        summary_sheet.append(["指标", "值"])
        summary_sheet.append(["处理类型", "通用表格合并"])
        summary_sheet.append(["模板", config.name])
        summary_sheet.append(["输入文件数", len(paths)])
        for sheet_name, count in sheet_row_counts.items():
            summary_sheet.append([f"{sheet_name}行数", count])

        if "来源映射" in target_workbook.sheetnames:
            del target_workbook["来源映射"]
        mapping_sheet = target_workbook.create_sheet("来源映射")
        mapping_sheet.append(["Sheet", "合并后行号", "来源文件", "来源行号"])
        for row in source_rows:
            mapping_sheet.append(row)

        target_workbook.save(output_file)
        return GenericMergeResult(
            output_file=output_file,
            summary={
                "处理类型": "通用表格合并",
                "模板": config.name,
                "输入文件数": len(paths),
                "各Sheet行数": sheet_row_counts,
                "输出文件": output_file.name,
            },
        )
    finally:
        target_workbook.close()
