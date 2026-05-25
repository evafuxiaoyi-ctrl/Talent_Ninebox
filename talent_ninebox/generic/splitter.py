from __future__ import annotations

import copy
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.formula import ArrayFormula

from .models import GenericSheetConfig, GenericSplitResult, GenericTemplateConfig

MAX_GENERIC_SPLIT_GROUPS = 120


def _clean(value: object) -> str:
    return str(value or "").strip()


def _safe_filename_part(value: object) -> str:
    text = _clean(value) or "未填写"
    text = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text[:60] or "未填写"


def _headers(sheet, header_row: int) -> list[str]:
    return [_clean(cell.value) for cell in sheet[header_row]]


def _field_column(headers: list[str], aliases: list[str]) -> int | None:
    normalized = [re.sub(r"\s+", "", header) for header in headers]
    for alias in aliases:
        target = re.sub(r"\s+", "", alias)
        for idx, header in enumerate(normalized, start=1):
            if header == target:
                return idx
    for alias in aliases:
        target = re.sub(r"\s+", "", alias)
        for idx, header in enumerate(normalized, start=1):
            if target and target in header:
                return idx
    return None


def _is_formula(value: object) -> bool:
    return (isinstance(value, str) and value.startswith("=")) or isinstance(value, ArrayFormula)


def _translate_formula(value: object, source, target) -> object:
    if isinstance(value, ArrayFormula):
        origin = value.ref or source.coordinate
        text = value.text
    else:
        origin = source.coordinate
        text = value
    if not isinstance(text, str):
        return value
    if source.coordinate != target.coordinate:
        try:
            text = Translator(text, origin=origin).translate_formula(target.coordinate)
        except Exception:
            pass
    if isinstance(value, ArrayFormula):
        return ArrayFormula(target.coordinate, text)
    return text


def _copy_cell(source, target) -> None:
    value = source.value
    if _is_formula(value):
        value = _translate_formula(value, source, target)
    target.value = value
    if source.has_style:
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)
    if source.hyperlink:
        target._hyperlink = copy.copy(source.hyperlink)
    if source.comment:
        target.comment = copy.copy(source.comment)


def _row_has_data(sheet, row: int, max_col: int) -> bool:
    return any(sheet.cell(row=row, column=col).value not in (None, "") for col in range(1, max_col + 1))


def _collect_groups(workbook_path: Path, config: GenericTemplateConfig, split_field: str) -> dict[str, dict[str, int]]:
    aliases = config.split_fields.get(split_field)
    if not aliases:
        raise ValueError(f"模板「{config.name}」不支持按「{split_field}」拆分。")

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        groups: dict[str, dict[str, int]] = {}
        for sheet_config in config.sheets:
            if sheet_config.name not in workbook.sheetnames:
                raise ValueError(f"未找到 Sheet「{sheet_config.name}」。")
            sheet = workbook[sheet_config.name]
            column = _field_column(_headers(sheet, sheet_config.header_row), aliases)
            if column is None:
                raise ValueError(f"Sheet「{sheet_config.name}」未找到拆分字段「{split_field}」。")
            for row in range(sheet_config.data_start_row, sheet.max_row + 1):
                if not _row_has_data(sheet, row, sheet.max_column):
                    continue
                value = _clean(sheet.cell(row=row, column=column).value) or "未填写"
                groups.setdefault(value, {}).setdefault(sheet_config.name, 0)
                groups[value][sheet_config.name] += 1
        if not groups:
            raise ValueError("未找到可拆分的数据行。")
        if len(groups) > MAX_GENERIC_SPLIT_GROUPS:
            raise ValueError(f"拆分后将生成 {len(groups)} 个文件，超过 {MAX_GENERIC_SPLIT_GROUPS} 个上限。")
        return groups
    finally:
        workbook.close()


def _filter_sheet(source_sheet, target_sheet, sheet_config: GenericSheetConfig, split_column: int, group_value: str) -> int:
    kept = 0
    for source_row in range(sheet_config.data_start_row, source_sheet.max_row + 1):
        if not _row_has_data(source_sheet, source_row, source_sheet.max_column):
            continue
        value = _clean(source_sheet.cell(row=source_row, column=split_column).value) or "未填写"
        if value != group_value:
            continue
        target_row = sheet_config.data_start_row + kept
        for column in range(1, source_sheet.max_column + 1):
            _copy_cell(source_sheet.cell(row=source_row, column=column), target_sheet.cell(row=target_row, column=column))
        if source_row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[target_row].height = source_sheet.row_dimensions[source_row].height
        kept += 1

    first_extra_row = sheet_config.data_start_row + kept
    if first_extra_row <= target_sheet.max_row:
        target_sheet.delete_rows(first_extra_row, target_sheet.max_row - first_extra_row + 1)
    return kept


def _write_group_workbook(source_path: Path, target_path: Path, config: GenericTemplateConfig, split_field: str, group_value: str) -> dict[str, int]:
    aliases = config.split_fields[split_field]
    source_workbook = load_workbook(source_path, data_only=False)
    target_workbook = load_workbook(target_path, data_only=False)
    try:
        counts: dict[str, int] = {}
        for sheet_config in config.sheets:
            source_sheet = source_workbook[sheet_config.name]
            target_sheet = target_workbook[sheet_config.name]
            split_column = _field_column(_headers(source_sheet, sheet_config.header_row), aliases)
            if split_column is None:
                raise ValueError(f"Sheet「{sheet_config.name}」未找到拆分字段「{split_field}」。")
            counts[sheet_config.name] = _filter_sheet(source_sheet, target_sheet, sheet_config, split_column, group_value)
        target_workbook.save(target_path)
        return counts
    finally:
        source_workbook.close()
        target_workbook.close()


def split_workbook_by_config(workbook_path: str | Path, output_dir: str | Path, config: GenericTemplateConfig, split_field: str) -> GenericSplitResult:
    workbook_path = Path(workbook_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    groups = _collect_groups(workbook_path, config, split_field)

    split_dir = output_dir / "split_excels"
    split_dir.mkdir(parents=True, exist_ok=True)
    generated: list[dict[str, Any]] = []
    for index, group_value in enumerate(sorted(groups.keys()), start=1):
        filename = f"{index:02d}_{_safe_filename_part(split_field)}_{_safe_filename_part(group_value)}.xlsx"
        target_path = split_dir / filename
        shutil.copy2(workbook_path, target_path)
        sheet_counts = _write_group_workbook(workbook_path, target_path, config, split_field, group_value)
        generated.append({"字段值": group_value, "文件名": filename, "各Sheet行数": sheet_counts})

    output_file = output_dir / f"{_safe_filename_part(config.name)}拆分结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    with zipfile.ZipFile(output_file, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for item in generated:
            archive.write(split_dir / str(item["文件名"]), arcname=str(item["文件名"]))

    return GenericSplitResult(
        output_file=output_file,
        summary={
            "处理类型": "通用表格拆分",
            "模板": config.name,
            "拆分依据": split_field,
            "生成文件数": len(generated),
            "拆分明细": generated,
        },
    )

