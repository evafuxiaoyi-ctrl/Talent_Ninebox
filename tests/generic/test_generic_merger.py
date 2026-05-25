from __future__ import annotations

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from talent_ninebox.generic.configs import ORG_INVENTORY_TEMPLATE
from talent_ninebox.generic.merger import merge_workbooks_by_config


def _make_org_inventory_workbook(path, prefix: str, first_dept: str) -> None:
    workbook = Workbook()
    org = workbook.active
    org.title = "组织架构盘点表"
    org.append(["名称", "编码", "一级组织", "二级组织", "三级组织", "负责人邮箱", "动作", "校验公式"])
    org.append([f"{prefix}中心", f"{prefix}001", first_dept, "销售中心", "", "a@example.com", "保留", '=A2&"-"&C2'])
    org.append([f"{prefix}分公司", f"{prefix}002", first_dept, "销售中心", f"{prefix}分公司", "b@example.com", "清理", '=A3&"-"&C3'])
    validation = DataValidation(type="list", formula1='"保留,清理"', allow_blank=True)
    validation.add("G2:G200")
    org.add_data_validation(validation)

    duty = workbook.create_sheet("组织职责盘点表")
    duty.append(["编码", "一级组织", "二级组织", "三级组织", "负责人邮箱", "组织职责"])
    duty.append([f"{prefix}001", first_dept, "销售中心", "", "a@example.com", f"{prefix}销售管理"])
    duty.append([f"{prefix}002", first_dept, "销售中心", f"{prefix}分公司", "b@example.com", f"{prefix}分公司销售"])
    workbook.save(path)


def test_generic_merge_combines_all_configured_sheets_and_tracks_sources(tmp_path) -> None:
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    _make_org_inventory_workbook(first, "A", "国内事业部")
    _make_org_inventory_workbook(second, "B", "技术事业部")

    result = merge_workbooks_by_config([first, second], tmp_path / "out", ORG_INVENTORY_TEMPLATE)

    assert result.output_file.exists()
    assert result.summary["各Sheet行数"] == {"组织架构盘点表": 4, "组织职责盘点表": 4}
    workbook = load_workbook(result.output_file, data_only=False)
    try:
        org = workbook["组织架构盘点表"]
        duty = workbook["组织职责盘点表"]
        assert org.max_row == 5
        assert duty.max_row == 5
        assert [org.cell(row=row, column=3).value for row in range(2, 6)] == ["国内事业部", "国内事业部", "技术事业部", "技术事业部"]
        assert [duty.cell(row=row, column=2).value for row in range(2, 6)] == ["国内事业部", "国内事业部", "技术事业部", "技术事业部"]
        assert org["H4"].value == '=A4&"-"&C4'
        assert org["I4"].value == "second.xlsx"
        assert org["J4"].value == "组织架构盘点表"
        assert org["K4"].value == 2
        assert len(org.data_validations.dataValidation) == 1
        assert str(org.data_validations.dataValidation[0].sqref) == "G2:G200"
        assert "处理摘要" in workbook.sheetnames
        assert "来源映射" in workbook.sheetnames
    finally:
        workbook.close()

