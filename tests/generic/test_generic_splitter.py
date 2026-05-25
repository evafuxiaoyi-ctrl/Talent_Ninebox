from __future__ import annotations

import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from talent_ninebox.generic.configs import ORG_INVENTORY_TEMPLATE
from talent_ninebox.generic.splitter import split_workbook_by_config


def _make_org_inventory_workbook(path) -> None:
    workbook = Workbook()
    org = workbook.active
    org.title = "组织架构盘点表"
    org.append(["名称", "编码", "一级组织", "二级组织", "三级组织", "负责人邮箱", "动作"])
    org.append(["销售中心", "D001", "国内事业部", "销售中心", "", "a@example.com", "保留"])
    org.append(["北京分公司", "D002", "国内事业部", "销售中心", "北京分公司", "b@example.com", "清理"])
    org.append(["技术中心", "D003", "技术事业部", "平台中心", "", "c@example.com", "保留"])
    validation = DataValidation(type="list", formula1='"保留,清理"', allow_blank=True)
    validation.add("G2:G200")
    org.add_data_validation(validation)

    duty = workbook.create_sheet("组织职责盘点表")
    duty.append(["编码", "一级组织", "二级组织", "三级组织", "负责人邮箱", "组织职责"])
    duty.append(["D001", "国内事业部", "销售中心", "", "a@example.com", "销售管理"])
    duty.append(["D002", "国内事业部", "销售中心", "北京分公司", "b@example.com", "北京销售"])
    duty.append(["D003", "技术事业部", "平台中心", "", "c@example.com", "平台研发"])
    workbook.save(path)


def test_generic_split_filters_all_configured_sheets_and_preserves_dropdown(tmp_path) -> None:
    workbook_path = tmp_path / "org.xlsx"
    output_dir = tmp_path / "out"
    _make_org_inventory_workbook(workbook_path)

    result = split_workbook_by_config(workbook_path, output_dir, ORG_INVENTORY_TEMPLATE, "一级部门")

    assert result.output_file.suffix == ".zip"
    assert result.summary["生成文件数"] == 2
    with zipfile.ZipFile(result.output_file) as archive:
        names = archive.namelist()
        assert any("国内事业部" in name for name in names)
        domestic_name = next(name for name in names if "国内事业部" in name)
        extracted = tmp_path / domestic_name
        extracted.write_bytes(archive.read(domestic_name))

    workbook = load_workbook(extracted, data_only=False)
    try:
        org = workbook["组织架构盘点表"]
        duty = workbook["组织职责盘点表"]
        assert org.max_row == 3
        assert duty.max_row == 3
        assert [org.cell(row=row, column=3).value for row in range(2, org.max_row + 1)] == ["国内事业部", "国内事业部"]
        assert [duty.cell(row=row, column=2).value for row in range(2, duty.max_row + 1)] == ["国内事业部", "国内事业部"]
        assert len(org.data_validations.dataValidation) == 1
        assert str(org.data_validations.dataValidation[0].sqref) == "G2:G200"
    finally:
        workbook.close()


def test_generic_split_can_use_owner_email_across_sheets(tmp_path) -> None:
    workbook_path = tmp_path / "org.xlsx"
    output_dir = tmp_path / "out"
    _make_org_inventory_workbook(workbook_path)

    result = split_workbook_by_config(workbook_path, output_dir, ORG_INVENTORY_TEMPLATE, "负责人邮箱")

    assert result.summary["生成文件数"] == 3
    with zipfile.ZipFile(result.output_file) as archive:
        names = archive.namelist()
        assert any("a@example.com" in name for name in names)
        owner_name = next(name for name in names if "a@example.com" in name)
        extracted = tmp_path / owner_name
        extracted.write_bytes(archive.read(owner_name))

    workbook = load_workbook(extracted, data_only=False)
    try:
        org = workbook["组织架构盘点表"]
        duty = workbook["组织职责盘点表"]
        assert org.max_row == 2
        assert duty.max_row == 2
        assert org["F2"].value == "a@example.com"
        assert duty["E2"].value == "a@example.com"
    finally:
        workbook.close()

