from __future__ import annotations

import zipfile
import shutil

from openpyxl import Workbook, load_workbook

from talent_ninebox.core.splitter import list_split_fields, list_split_sheets, split_workbook_by_field


def _make_split_workbook(path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "人才盘点"
    sheet.append(["说明"])
    sheet.append(["请填写基础字段"])
    sheet.append(["员工姓名", "工号", "邮箱", "岗位", "盘点人员类型", "一级部门", "二级部门", "盘点人", "备注公式"])
    sheet.append(["示例行", "000", "sample@example.com", "示例", "XXX", "XXX", "XXX", "李四", '=A4&"-"&F4'])
    sheet.append(["张三", "001", "a@example.com", "顾问", "业务负责人", "销售部", "华东", "傅肖忆", '=A5&"-"&F5'])
    sheet.append(["李四", "002", "b@example.com", "主管", "关键人才", "教研部", "数学", "赵桐", '=A6&"-"&F6'])
    sheet.append(["王五", "003", "c@example.com", "顾问", "业务负责人", "销售部", "华南", "傅肖忆", None])
    params = workbook.create_sheet("参数")
    params["A1"] = "保留"
    params["B3"] = "盘点人员类型"
    workbook.save(path)


def _inject_x14_data_validation(path) -> None:
    ext_xml = (
        '<extLst><ext uri="{CCE6A557-97BC-4b89-ADB6-D9C93CAAB3DF}" '
        'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" '
        'xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main">'
        '<x14:dataValidations count="1">'
        '<x14:dataValidation type="list" allowBlank="1" showErrorMessage="1" '
        'xr:uid="{11111111-1111-1111-1111-111111111111}">'
        '<x14:formula1><xm:f>参数!$B$3:$B$4</xm:f></x14:formula1>'
        '<xm:sqref>E4:E100</xm:sqref>'
        '</x14:dataValidation>'
        '</x14:dataValidations></ext></extLst>'
    )
    tmp_path = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as target:
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                xml = data.decode("utf-8")
                xml = xml.replace(
                    '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
                    '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                    'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
                    'mc:Ignorable="x14 xr" '
                    'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision">',
                    1,
                )
                xml = xml.replace("</worksheet>", f"{ext_xml}</worksheet>", 1)
                data = xml.encode("utf-8")
            target.writestr(info, data)
    shutil.move(tmp_path, path)


def test_list_split_fields_reads_allowed_headers_from_row_3(tmp_path) -> None:
    workbook_path = tmp_path / "template.xlsx"
    _make_split_workbook(workbook_path)

    fields = list_split_fields(workbook_path)

    assert [field.name for field in fields] == ["盘点人员类型", "一级部门", "二级部门", "盘点人"]
    assert fields[0].column_letter == "E"


def test_list_split_sheets_returns_sheet_fields(tmp_path) -> None:
    workbook_path = tmp_path / "template.xlsx"
    _make_split_workbook(workbook_path)

    sheets = list_split_sheets(workbook_path)

    assert len(sheets) == 1
    assert sheets[0].name == "人才盘点"
    assert [field.name for field in sheets[0].fields] == ["盘点人员类型", "一级部门", "二级部门", "盘点人"]


def test_split_workbook_by_field_creates_zip_with_grouped_excels(tmp_path) -> None:
    workbook_path = tmp_path / "template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    _make_split_workbook(workbook_path)

    result = split_workbook_by_field(workbook_path, output_dir, "一级部门", "人才盘点")

    assert result.output_file.suffix == ".zip"
    assert result.summary["生成文件数"] == 2
    assert result.summary["总人员数"] == 3

    with zipfile.ZipFile(result.output_file) as archive:
        names = archive.namelist()
        assert len(names) == 2
        assert any("销售部" in name for name in names)
        archive.extractall(tmp_path / "unzipped")

    sales_file = next((tmp_path / "unzipped").glob("*销售部*.xlsx"))
    workbook = load_workbook(sales_file, data_only=False)
    try:
        sheet = workbook["人才盘点"]
        assert sheet.max_row == 5
        assert sheet["A4"].value == "张三"
        assert sheet["A5"].value == "王五"
        assert sheet["I5"].value == '=A5&"-"&F5'
        assert "参数" in workbook.sheetnames
    finally:
        workbook.close()


def test_split_workbook_preserves_x14_data_validation_extensions(tmp_path) -> None:
    workbook_path = tmp_path / "template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    _make_split_workbook(workbook_path)
    _inject_x14_data_validation(workbook_path)

    result = split_workbook_by_field(workbook_path, output_dir, "一级部门", "人才盘点")

    with zipfile.ZipFile(result.output_file) as archive:
        sales_name = next(name for name in archive.namelist() if "销售部" in name)
        sales_file = tmp_path / sales_name
        sales_file.write_bytes(archive.read(sales_name))

    with zipfile.ZipFile(sales_file) as workbook_archive:
        sheet_xml = workbook_archive.read("xl/worksheets/sheet1.xml").decode("utf-8")

    assert "x14:dataValidations" in sheet_xml
    assert '<dataValidations count="1">' in sheet_xml
    assert '<dataValidation type="list" allowBlank="1" showErrorMessage="1" sqref="E4:E100">' in sheet_xml
    assert "参数!$B$3:$B$4" in sheet_xml
    assert "E4:E100" in sheet_xml
