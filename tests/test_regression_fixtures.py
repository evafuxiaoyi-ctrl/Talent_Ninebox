from __future__ import annotations

from openpyxl import load_workbook

from scripts.generate_regression_fixtures import generate_fixture_set
from talent_ninebox.core.models import ProcessOptions
from talent_ninebox.core.processor import process_workbook, process_zip


def test_generated_regression_fixtures_cover_initial_and_final(tmp_path) -> None:
    fixtures = generate_fixture_set(tmp_path / "fixtures")

    initial = process_zip(fixtures["initial_10"], tmp_path / "initial_out", ProcessOptions(placement_mode="initial", stage_label="初版整合"))
    assert initial.summary.excel_file_count == 10
    assert initial.summary.processed_file_count == 10
    assert initial.summary.total_people == 30
    assert initial.summary.placed_people == 30

    max_batch = process_zip(fixtures["initial_30"], tmp_path / "max_out", ProcessOptions(placement_mode="initial", stage_label="初版整合"))
    assert max_batch.summary.excel_file_count == 30
    assert max_batch.summary.total_people == 90

    missing = process_zip(fixtures["missing_field"], tmp_path / "missing_out", ProcessOptions(placement_mode="initial", stage_label="初版整合"))
    assert missing.summary.excel_file_count == 3
    assert missing.summary.skipped_file_count == 1
    assert any(issue.category == "模板异常" for issue in missing.issues)

    blank = process_zip(fixtures["blank_placement"], tmp_path / "blank_out", ProcessOptions(placement_mode="initial", stage_label="初版整合"))
    assert blank.summary.unplaced_people >= 1

    final = process_workbook(fixtures["final_workbook"], tmp_path / "final_out", ProcessOptions(placement_mode="final", stage_label="终版九宫格生成", add_source_columns=False))
    assert final.summary.total_people == 30
    assert final.summary.placed_people == 30

    wb = load_workbook(final.output_file, data_only=False)
    assert "人才九宫格" in wb.sheetnames
    assert "处理摘要" in wb.sheetnames
