from __future__ import annotations

import io
import zipfile

from fastapi.testclient import TestClient
from openpyxl import Workbook

from talent_ninebox.web import app as web_app
from talent_ninebox.web.app import app


def test_login_required() -> None:
    client = TestClient(app)
    response = client.get("/", follow_redirects=False)
    assert response.status_code == 200
    assert "进入" in response.text


def test_dingtalk_login_hides_password_form(monkeypatch) -> None:
    monkeypatch.setenv("DINGTALK_CORP_ID", "ding-test")
    monkeypatch.setenv("DINGTALK_APP_KEY", "ding-app")
    monkeypatch.setenv("DINGTALK_APP_SECRET", "secret")

    client = TestClient(app)
    response = client.get("/", follow_redirects=False)

    assert response.status_code == 200
    assert "正在获取钉钉身份" in response.text
    assert "访问密码" not in response.text
    assert 'type="password"' not in response.text


def test_health_check() -> None:
    client = TestClient(app)
    response = client.get("/health")
    assert response.status_code == 200
    assert response.json() == {"status": "ok"}


def test_login_success(monkeypatch) -> None:
    monkeypatch.setenv("APP_ACCESS_PASSWORD", "secret")
    client = TestClient(app)
    response = client.post("/login", data={"password": "secret"}, follow_redirects=False)
    assert response.status_code == 200
    assert "初版整合" in response.text


def test_download_falls_back_to_tmp_file(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APP_ACCESS_PASSWORD", "secret")
    monkeypatch.setattr(web_app, "TMP_ROOT", tmp_path)
    web_app.TASKS.clear()

    task_id = "a" * 32
    output_dir = tmp_path / task_id / "output"
    output_dir.mkdir(parents=True)
    output_file = output_dir / "人才盘点整合结果_20260512.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "ok"
    workbook.save(output_file)

    client = TestClient(app)
    client.post("/login", data={"password": "secret"})
    response = client.get(f"/download/{task_id}")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_process_file_hides_unexpected_error_details(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APP_ACCESS_PASSWORD", "secret")
    monkeypatch.setattr(web_app, "TMP_ROOT", tmp_path)
    web_app.TASKS.clear()

    async def fail_processing(mode, file):
        raise RuntimeError("sensitive workbook path and stack details")

    monkeypatch.setattr(web_app, "_run_processing", fail_processing)

    client = TestClient(app)
    client.post("/login", data={"password": "secret"})
    response = client.post(
        "/process-file",
        data={"mode": "initial"},
        files={"file": ("upload.zip", io.BytesIO(b"not a real zip"), "application/zip")},
    )

    assert response.status_code == 400
    assert "sensitive" not in response.text
    assert response.json()["error"].startswith("处理失败")


def test_process_file_returns_guidance_for_wrong_mode_upload(monkeypatch) -> None:
    monkeypatch.setenv("APP_ACCESS_PASSWORD", "secret")
    client = TestClient(app)
    client.post("/login", data={"password": "secret"})
    response = client.post(
        "/tasks",
        data={"mode": "final"},
        files={"file": ("upload.zip", io.BytesIO(b"fake"), "application/zip")},
    )

    assert response.status_code == 400
    payload = response.json()
    assert payload["error"] == "终版生成仅支持上传已整合后的 .xlsx 文件。"
    assert any(".xlsx" in item for item in payload["guidance"])


def test_task_flow_returns_download_url(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APP_ACCESS_PASSWORD", "secret")
    monkeypatch.setattr(web_app, "TMP_ROOT", tmp_path)
    web_app.TASKS.clear()

    workbook_path = tmp_path / "dept.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "盘点"
    sheet.append(["工号", "姓名", "初始九宫格落位"])
    sheet.append(["001", "张三", "高潜高绩"])
    workbook.save(workbook_path)

    zip_bytes = io.BytesIO()
    with zipfile.ZipFile(zip_bytes, "w") as archive:
        archive.write(workbook_path, "dept.xlsx")
    zip_bytes.seek(0)

    client = TestClient(app)
    client.post("/login", data={"password": "secret"})
    response = client.post(
        "/tasks",
        data={"mode": "initial"},
        files={"file": ("upload.zip", zip_bytes, "application/zip")},
    )

    assert response.status_code == 202
    task_id = response.json()["task_id"]
    status_response = client.get(f"/tasks/{task_id}")
    payload = status_response.json()
    assert payload["status"] == "done"
    assert payload["download_url"] == f"/download/{task_id}"
    assert payload["summary"]["总人员数"] == 1


def test_split_fields_and_task_flow(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APP_ACCESS_PASSWORD", "secret")
    monkeypatch.setattr(web_app, "TMP_ROOT", tmp_path)
    web_app.TASKS.clear()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "人才盘点"
    sheet.append(["说明"])
    sheet.append(["填写说明"])
    sheet.append(["员工姓名", "工号", "盘点人员类型", "一级部门", "二级部门", "盘点人"])
    sheet.append(["张三", "001", "业务负责人", "销售部", "华东", "傅肖忆"])
    sheet.append(["李四", "002", "关键人才", "教研部", "数学", "赵桐"])
    workbook_bytes = io.BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)

    client = TestClient(app)
    client.post("/login", data={"password": "secret"})

    fields_response = client.post(
        "/split-fields",
        files={"file": ("template.xlsx", io.BytesIO(workbook_bytes.getvalue()), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert fields_response.status_code == 200
    sheets = fields_response.json()["sheets"]
    assert sheets[0]["name"] == "人才盘点"
    assert [field["name"] for field in sheets[0]["fields"]] == ["盘点人员类型", "一级部门", "二级部门", "盘点人"]

    task_response = client.post(
        "/tasks",
        data={"mode": "split", "split_sheet": "人才盘点", "split_field": "一级部门"},
        files={"file": ("template.xlsx", io.BytesIO(workbook_bytes.getvalue()), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert task_response.status_code == 202
    task_id = task_response.json()["task_id"]
    status_response = client.get(f"/tasks/{task_id}")
    payload = status_response.json()
    assert payload["status"] == "done"
    assert payload["summary"]["处理类型"] == "表格拆分"
    assert payload["summary"]["生成文件数"] == 2

    download_response = client.get(f"/download/{task_id}")
    assert download_response.status_code == 200
    assert download_response.headers["content-type"] == "application/zip"


def _make_org_workbook_bytes(prefix: str = "A") -> bytes:
    workbook = Workbook()
    org = workbook.active
    org.title = "组织架构盘点表"
    org.append(["名称", "编码", "一级组织", "二级组织", "三级组织", "负责人邮箱", "动作"])
    org.append([f"{prefix}中心", f"{prefix}001", "国内事业部", "销售中心", "", "a@example.com", "保留"])
    org.append([f"{prefix}平台", f"{prefix}002", "技术事业部", "平台中心", "", "b@example.com", "保留"])
    duty = workbook.create_sheet("组织职责盘点表")
    duty.append(["编码", "一级组织", "二级组织", "三级组织", "负责人邮箱", "组织职责"])
    duty.append([f"{prefix}001", "国内事业部", "销售中心", "", "a@example.com", "销售管理"])
    duty.append([f"{prefix}002", "技术事业部", "平台中心", "", "b@example.com", "平台研发"])
    workbook_bytes = io.BytesIO()
    workbook.save(workbook_bytes)
    return workbook_bytes.getvalue()


def test_org_split_task_flow(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APP_ACCESS_PASSWORD", "secret")
    monkeypatch.setattr(web_app, "TMP_ROOT", tmp_path)
    web_app.TASKS.clear()

    client = TestClient(app)
    client.post("/login", data={"password": "secret"})
    response = client.post(
        "/tasks",
        data={"mode": "org_split", "split_field": "一级部门"},
        files={"file": ("org.xlsx", io.BytesIO(_make_org_workbook_bytes()), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 202
    task_id = response.json()["task_id"]
    payload = client.get(f"/tasks/{task_id}").json()
    assert payload["status"] == "done"
    assert payload["summary"]["处理类型"] == "通用表格拆分"
    assert payload["summary"]["生成文件数"] == 2


def test_org_merge_task_flow(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APP_ACCESS_PASSWORD", "secret")
    monkeypatch.setattr(web_app, "TMP_ROOT", tmp_path)
    web_app.TASKS.clear()

    zip_bytes = io.BytesIO()
    with zipfile.ZipFile(zip_bytes, "w") as archive:
        archive.writestr("first.xlsx", _make_org_workbook_bytes("A"))
        archive.writestr("second.xlsx", _make_org_workbook_bytes("B"))
    zip_bytes.seek(0)

    client = TestClient(app)
    client.post("/login", data={"password": "secret"})
    response = client.post(
        "/tasks",
        data={"mode": "org_merge"},
        files={"file": ("orgs.zip", zip_bytes, "application/zip")},
    )

    assert response.status_code == 202
    task_id = response.json()["task_id"]
    payload = client.get(f"/tasks/{task_id}").json()
    assert payload["status"] == "done"
    assert payload["summary"]["处理类型"] == "通用表格合并"
    assert payload["summary"]["输入文件数"] == 2
    assert payload["summary"]["各Sheet行数"] == {"组织架构盘点表": 4, "组织职责盘点表": 4}
