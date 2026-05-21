from __future__ import annotations

import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from talent_ninebox.core.models import ProcessOptions, RowRecord
from talent_ninebox.core.processor import _build_ninebox, process_zip
from talent_ninebox.core.ninebox_mapper import NINEBOX_KEYS


def _make_workbook(path: Path, rows: list[tuple[str, str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "人才盘点数据收集表"
    ws.append(["说明", "", "", ""])
    ws.append(["工号", "姓名", "所属部门", "人才九宫格落位", "评分"])
    for idx, row in enumerate(rows, start=3):
        emp_id, name, dept, place = row
        ws.append([emp_id, name, dept, place, f"=LEN(B{idx})"])
    ws.column_dimensions["B"].width = 18
    wb.save(path)


def test_process_zip_generates_output(tmp_path: Path) -> None:
    first = tmp_path / "first.xlsx"
    second = tmp_path / "nested" / "second.xlsx"
    second.parent.mkdir()
    _make_workbook(first, [("001", "张三", "产品部", "高潜高绩"), ("002", "李四", "产品部", "B2")])
    _make_workbook(second, [("003", "王五", "研发部", "低潜低绩"), ("004", "赵六", "研发部", "未知落位")])

    zip_path = tmp_path / "input.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.write(first, "first.xlsx")
        zf.write(second, "nested/second.xlsx")
        zf.writestr("__MACOSX/ignored.xlsx", "ignored")
        zf.writestr("~$temp.xlsx", "ignored")

    result = process_zip(zip_path, tmp_path / "out")

    assert result.output_file.exists()
    assert result.summary.excel_file_count == 2
    assert result.summary.processed_file_count == 2
    assert result.summary.total_people == 4
    assert result.summary.placed_people == 3
    assert result.summary.unplaced_people == 1

    wb = load_workbook(result.output_file, data_only=False)
    assert set(["使用说明", "处理摘要", "人才盘点数据收集表", "人才九宫格", "异常报告", "文件来源映射"]).issubset(wb.sheetnames)
    assert "整合总表" not in wb.sheetnames
    merged = wb["人才盘点数据收集表"]
    assert merged.max_row == 6
    assert merged["A1"].value == "说明"
    assert merged["E3"].value == "=LEN(B3)"
    assert merged["E4"].value == "=LEN(B4)"
    assert merged["B3"].value == "张三"
    assert merged["B6"].value == "赵六"

    assert list(result.summary.ninebox_counts) == NINEBOX_KEYS
    assert NINEBOX_KEYS == [
        "高潜力 / 低绩效",
        "高潜力 / 中绩效",
        "高潜力 / 高绩效",
        "中潜力 / 低绩效",
        "中潜力 / 中绩效",
        "中潜力 / 高绩效",
        "低潜力 / 低绩效",
        "低潜力 / 中绩效",
        "低潜力 / 高绩效",
    ]

    summary = wb["处理摘要"]
    rows = list(summary.iter_rows(values_only=True))
    ninebox_row_index = next(idx for idx, row in enumerate(rows) if row[0] == "九宫格人数")
    assert rows[ninebox_row_index + 1] == ("  高潜力 / 低绩效", 0, 0)
    assert rows[ninebox_row_index + 3] == ("  高潜力 / 高绩效", 1, 0.25)

    ninebox = wb["人才九宫格"]
    assert ninebox["A3"].value == "能力：高"
    assert ninebox["B3"].value == "高能力 / 低绩效（0人，0%）"
    assert ninebox["J3"].value == "高能力 / 高绩效（1人，25%）"


def test_ninebox_uses_cached_formula_placement_before_deriving() -> None:
    headers = ["工号", "姓名", "初始九宫格落位", "25+26绩效综合等级", "能力九宫格落位"]
    records = [
        RowRecord(
            ["001", "公式人员1", "=AW3&AC3", "B", "=formula"],
            "source.xlsx",
            "人才盘点数据收集表",
            3,
            calculated_values=["001", "公式人员1", "高绩效高能力", "B", "低"],
        ),
        RowRecord(
            ["002", "公式人员2", "=AW4&AC4", "B", "=formula"],
            "source.xlsx",
            "人才盘点数据收集表",
            4,
            calculated_values=["002", "公式人员2", "高绩效高能力", "B", "低"],
        ),
    ]

    people_by_box = _build_ninebox(records, headers, [], ProcessOptions(placement_mode="initial"))

    assert len(people_by_box["高潜力 / 高绩效"]) == 2


def test_process_zip_preserves_template_main_sheet_for_support_formulas(tmp_path: Path) -> None:
    workbook_path = tmp_path / "template_with_support.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "人才盘点数据收集表"
    ws.append(["说明", "", "", ""])
    ws.append(["工号", "姓名", "所属部门", "人才九宫格落位", "评分"])
    ws.append(["001", "张三", "产品部", "高潜高绩", "=LEN(B3)"])
    support = wb.create_sheet("供参考-九宫格校准页")
    support["A1"] = "=COUNTA(人才盘点数据收集表!$B$3:$B$5000)"
    wb.save(workbook_path)

    zip_path = tmp_path / "input.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.write(workbook_path, "template_with_support.xlsx")

    result = process_zip(zip_path, tmp_path / "out")

    output = load_workbook(result.output_file, data_only=False)
    try:
        assert "人才盘点数据收集表" in output.sheetnames
        assert "整合总表" not in output.sheetnames
        main = output["人才盘点数据收集表"]
        assert main["B3"].value == "张三"
        assert main["E3"].value == "=LEN(B3)"
        assert output["供参考-九宫格校准页"]["A1"].value == "=COUNTA(人才盘点数据收集表!$B$3:$B$5000)"
    finally:
        output.close()


def test_process_zip_skips_sample_row_style_for_merged_data(tmp_path: Path) -> None:
    workbook_path = tmp_path / "with_sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "人才盘点数据收集表"
    ws.append(["说明", "", "", ""])
    ws.append(["工号", "姓名", "所属部门", "人才九宫格落位", "评分"])
    ws.append(["示例行", "样例", "样例部门", "高潜高绩", "=LEN(B3)"])
    for cell in ws[3]:
        cell.font = Font(color="C00000")
    ws.append(["001", "张三", "产品部", "高潜高绩", "=LEN(B4)"])
    workbook_path.parent.mkdir(exist_ok=True)
    wb.save(workbook_path)

    zip_path = tmp_path / "input.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.write(workbook_path, "with_sample.xlsx")

    result = process_zip(zip_path, tmp_path / "out")

    assert result.summary.total_people == 1
    output = load_workbook(result.output_file, data_only=False)
    try:
        merged = output["人才盘点数据收集表"]
        assert merged["B3"].value == "张三"
        assert merged["B3"].font.color is None or merged["B3"].font.color.rgb != "00C00000"
    finally:
        output.close()


def test_process_zip_highlights_value_score_from_value_dimension_text(tmp_path: Path) -> None:
    workbook_path = tmp_path / "value_score.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "人才盘点数据收集表"
    ws.append(["说明", "", "", "", "", "", "", "", ""])
    ws.append(["工号", "姓名", "人才九宫格落位", "成就客户", "激情人生", "拥抱变化", "团队协作", "价值观综合得分", "离职风险"])
    ws.append(["001", "高分人员", "高潜高绩", "4、优秀", "4、优秀", "4、优秀", "4、优秀", "=AVERAGE(D3:G3)", "低"])
    ws.append(["002", "低分人员", "低潜低绩", "2、待提升", "2、待提升", "2、待提升", "2、待提升", "=AVERAGE(D4:G4)", "低"])
    ws.append(["003", "高风险人员", "中潜中绩", "3、符合", "3、符合", "3、符合", "3、符合", "=AVERAGE(D5:G5)", "高"])
    wb.save(workbook_path)

    zip_path = tmp_path / "input.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.write(workbook_path, "value_score.xlsx")

    result = process_zip(zip_path, tmp_path / "out")
    output = load_workbook(result.output_file, data_only=False)
    try:
        merged = output["人才盘点数据收集表"]
        assert merged["H3"].fill.fgColor.rgb == "0000B050"
        assert merged["H3"].font.color.rgb == "00FFFFFF"
        assert merged["H4"].fill.fgColor.rgb == "00FF0000"
        assert merged["H4"].font.color.rgb == "00FFFFFF"
        assert merged["I5"].fill.fgColor.rgb == "00FF0000"
        assert merged["I5"].font.color.rgb == "00FFFFFF"
    finally:
        output.close()


def test_process_zip_repairs_mojibake_source_filename(tmp_path: Path) -> None:
    workbook_path = tmp_path / "source.xlsx"
    _make_workbook(workbook_path, [("001", "张三", "产品部", "高潜高绩")])

    readable_name = "部门一_人才盘点.xlsx"
    mojibake_name = readable_name.encode("utf-8").decode("cp437")
    zip_path = tmp_path / "input.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.write(workbook_path, mojibake_name)

    result = process_zip(zip_path, tmp_path / "out")

    output = load_workbook(result.output_file, data_only=False)
    try:
        merged = output["人才盘点数据收集表"]
        assert merged["F3"].value == readable_name
        mapping = output["文件来源映射"]
        assert mapping["D2"].value == readable_name
    finally:
        output.close()


def test_too_many_excels_stops(tmp_path: Path) -> None:
    zip_path = tmp_path / "input.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        for idx in range(31):
            zf.writestr(f"{idx}.xlsx", "not real")

    try:
        process_zip(zip_path, tmp_path / "out")
    except ValueError as exc:
        assert "超过 30 个 Excel" in str(exc)
    else:
        raise AssertionError("expected ValueError")
